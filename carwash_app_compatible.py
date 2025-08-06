#!/usr/bin/env python3
"""
Carwash Management Desktop Application - Python 3.13 Compatible Version
A standalone desktop app for managing carwash operations with car tracking and payment processing.
Fixed compatibility issues with newer Python versions.
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import os
import csv
import json
from datetime import datetime
from pathlib import Path
import shutil
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

class CarwashApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Carwash Management System")
        self.root.geometry("1200x800")
        
        # Data storage
        self.cars_data = {}
        self.current_employee = None
        self.data_file = "carwash_data.json"
        
        # Car statuses
        self.STATUS_WASHING = "washing"
        self.STATUS_AWAITING_PAYMENT = "awaiting_payment"
        self.STATUS_FINISHED = "finished"
        
        # Load existing data
        self.load_data()
        
        # Create UI
        self.create_login_screen()
        
        # Auto-save every 30 seconds
        self.auto_save()
    
    def load_data(self):
        """Load car data from JSON file"""
        if os.path.exists(self.data_file):
            try:
                with open(self.data_file, 'r') as f:
                    data = json.load(f)
                    self.cars_data = data.get('cars', {})
                    # Convert timestamp strings back to datetime objects
                    for car_id, car_data in self.cars_data.items():
                        if 'timestamp' in car_data:
                            car_data['timestamp'] = datetime.fromisoformat(car_data['timestamp'])
                        if 'completion_time' in car_data and car_data['completion_time']:
                            car_data['completion_time'] = datetime.fromisoformat(car_data['completion_time'])
            except Exception as e:
                messagebox.showerror("Error", f"Failed to load data: {e}")
    
    def save_data(self):
        """Save car data to JSON file"""
        try:
            # Convert datetime objects to strings for JSON serialization
            data_to_save = {'cars': {}}
            for car_id, car_data in self.cars_data.items():
                car_copy = car_data.copy()
                if 'timestamp' in car_copy:
                    car_copy['timestamp'] = car_copy['timestamp'].isoformat()
                if 'completion_time' in car_copy and car_copy['completion_time']:
                    car_copy['completion_time'] = car_copy['completion_time'].isoformat()
                data_to_save['cars'][car_id] = car_copy
            
            with open(self.data_file, 'w') as f:
                json.dump(data_to_save, f, indent=2)
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save data: {e}")
    
    def auto_save(self):
        """Auto-save data every 30 seconds"""
        self.save_data()
        self.root.after(30000, self.auto_save)  # 30 seconds
    
    def create_login_screen(self):
        """Create the login interface"""
        self.clear_screen()
        
        # Configure root window style
        self.root.configure(bg='#f0f0f0')
        
        # Main frame
        login_frame = ttk.Frame(self.root, padding="50")
        login_frame.pack(expand=True, fill='both')
        
        # Title with custom styling
        title_frame = ttk.Frame(login_frame)
        title_frame.pack(pady=30)
        
        title_label = tk.Label(title_frame, text="Carwash Management System", 
                              font=('Arial', 24, 'bold'), 
                              bg='#f0f0f0', fg='#2c3e50')
        title_label.pack()
        
        # Employee name
        name_label = tk.Label(login_frame, text="Employee Name:", 
                             font=('Arial', 12), bg='#f0f0f0', fg='#34495e')
        name_label.pack(pady=5)
        
        self.name_entry = ttk.Entry(login_frame, width=30, font=('Arial', 12))
        self.name_entry.pack(pady=5)
        
        # Role selection
        role_label = tk.Label(login_frame, text="Role:", 
                             font=('Arial', 12), bg='#f0f0f0', fg='#34495e')
        role_label.pack(pady=(20, 5))
        
        self.role_var = tk.StringVar(value="washer")
        
        role_frame = ttk.Frame(login_frame)
        role_frame.pack(pady=5)
        
        # Use regular radiobuttons to avoid ttk font issues
        washer_radio = tk.Radiobutton(role_frame, text="Washer", variable=self.role_var, 
                                     value="washer", font=('Arial', 12),
                                     bg='#f0f0f0', fg='#34495e', selectcolor='#3498db')
        washer_radio.pack(side='left', padx=20)
        
        cashier_radio = tk.Radiobutton(role_frame, text="Cashier", variable=self.role_var, 
                                      value="cashier", font=('Arial', 12),
                                      bg='#f0f0f0', fg='#34495e', selectcolor='#3498db')
        cashier_radio.pack(side='left', padx=20)
        
        # Login button
        login_btn = ttk.Button(login_frame, text="Start Work", command=self.login)
        login_btn.pack(pady=30)
        
        # Bind Enter key to login
        self.root.bind('<Return>', lambda e: self.login())
        self.name_entry.focus()
    
    def login(self):
        """Handle employee login"""
        name = self.name_entry.get().strip()
        if not name:
            messagebox.showerror("Error", "Please enter your name")
            return
        
        self.current_employee = {
            'name': name,
            'role': self.role_var.get()
        }
        
        self.create_dashboard()
    
    def create_dashboard(self):
        """Create the main dashboard interface"""
        self.clear_screen()
        
        # Reset background
        self.root.configure(bg='SystemButtonFace')
        
        # Main container
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.pack(expand=True, fill='both')
        
        # Header
        header_frame = ttk.Frame(main_frame)
        header_frame.pack(fill='x', pady=(0, 10))
        
        header_text = f"Dashboard - {self.current_employee['name']} ({self.current_employee['role'].title()})"
        header_label = tk.Label(header_frame, text=header_text, 
                               font=('Arial', 16, 'bold'), fg='#2c3e50')
        header_label.pack(side='left')
        
        # Buttons frame
        btn_frame = ttk.Frame(header_frame)
        btn_frame.pack(side='right')
        
        if self.current_employee['role'] == 'washer':
            ttk.Button(btn_frame, text="Add New Car", 
                      command=self.show_add_car_dialog).pack(side='left', padx=5)
        
        ttk.Button(btn_frame, text="Export Excel", 
                  command=self.export_excel).pack(side='left', padx=5)
        ttk.Button(btn_frame, text="Export CSV", 
                  command=self.export_csv).pack(side='left', padx=5)
        ttk.Button(btn_frame, text="Reset Daily Data", 
                  command=self.reset_daily_data).pack(side='left', padx=5)
        ttk.Button(btn_frame, text="Logout", 
                  command=self.logout).pack(side='left', padx=5)
        
        # Status sections
        notebook = ttk.Notebook(main_frame)
        notebook.pack(expand=True, fill='both', pady=10)
        
        # Washing tab
        washing_frame = ttk.Frame(notebook)
        notebook.add(washing_frame, text="Washing")
        self.create_car_list(washing_frame, self.STATUS_WASHING, "Move to Payment")
        
        # Awaiting Payment tab
        payment_frame = ttk.Frame(notebook)
        notebook.add(payment_frame, text="Awaiting Payment")
        self.create_car_list(payment_frame, self.STATUS_AWAITING_PAYMENT, "Process Payment")
        
        # Finished tab
        finished_frame = ttk.Frame(notebook)
        notebook.add(finished_frame, text="Finished")
        self.create_car_list(finished_frame, self.STATUS_FINISHED, None)
        
        # Status bar
        self.status_var = tk.StringVar()
        self.update_status()
        status_bar = tk.Label(main_frame, textvariable=self.status_var, 
                             relief='sunken', font=('Arial', 10),
                             anchor='w', padx=10)
        status_bar.pack(fill='x', side='bottom')
        
        # Auto-refresh every 5 seconds
        self.auto_refresh()
    
    def create_car_list(self, parent, status, action_text):
        """Create a list of cars for a specific status"""
        # Create treeview
        columns = ('Car Name', 'Plate Number', 'Washer', 'Cashier', 'Payment', 'Time')
        tree = ttk.Treeview(parent, columns=columns, show='headings', height=15)
        
        # Configure columns
        for col in columns:
            tree.heading(col, text=col)
            tree.column(col, width=120)
        
        # Scrollbar
        scrollbar = ttk.Scrollbar(parent, orient='vertical', command=tree.yview)
        tree.configure(yscrollcommand=scrollbar.set)
        
        # Pack treeview and scrollbar
        tree.pack(side='left', expand=True, fill='both')
        scrollbar.pack(side='right', fill='y')
        
        # Populate with data
        self.populate_car_list(tree, status)
        
        # Action button frame
        if action_text:
            action_frame = ttk.Frame(parent)
            action_frame.pack(fill='x', pady=5)
            
            ttk.Button(action_frame, text=action_text, 
                      command=lambda: self.handle_action(tree, status)).pack()
        
        # Store reference for updates
        setattr(self, f'{status}_tree', tree)
    
    def populate_car_list(self, tree, status):
        """Populate the car list with current data"""
        # Clear existing items
        for item in tree.get_children():
            tree.delete(item)
        
        # Add cars with matching status
        for car_id, car_data in self.cars_data.items():
            if car_data['status'] == status:
                payment_text = f"₱{car_data.get('payment_amount', 0):.2f}" if car_data.get('payment_amount') else ""
                time_text = car_data['timestamp'].strftime('%H:%M:%S')
                
                tree.insert('', 'end', values=(
                    car_data['car_name'],
                    car_data['plate_number'],
                    car_data.get('washer_name', ''),
                    car_data.get('cashier_name', ''),
                    payment_text,
                    time_text
                ), tags=(car_id,))
    
    def handle_action(self, tree, status):
        """Handle action button clicks"""
        selection = tree.selection()
        if not selection:
            messagebox.showwarning("Warning", "Please select a car first")
            return
        
        car_id = tree.item(selection[0])['tags'][0]
        
        if status == self.STATUS_WASHING:
            self.move_to_payment(car_id)
        elif status == self.STATUS_AWAITING_PAYMENT:
            self.process_payment(car_id)
    
    def move_to_payment(self, car_id):
        """Move car from washing to awaiting payment"""
        if car_id in self.cars_data:
            self.cars_data[car_id]['status'] = self.STATUS_AWAITING_PAYMENT
            self.cars_data[car_id]['washer_name'] = self.current_employee['name']
            self.refresh_dashboard()
            messagebox.showinfo("Success", "Car moved to payment queue")
    
    def process_payment(self, car_id):
        """Process payment for a car"""
        if car_id not in self.cars_data:
            return
        
        # Create payment dialog
        dialog = tk.Toplevel(self.root)
        dialog.title("Process Payment")
        dialog.geometry("400x250")
        dialog.transient(self.root)
        dialog.grab_set()
        dialog.configure(bg='#f0f0f0')
        
        car_data = self.cars_data[car_id]
        
        # Car info with better styling
        car_label = tk.Label(dialog, text=f"Car: {car_data['car_name']}", 
                            font=('Arial', 14, 'bold'), bg='#f0f0f0', fg='#2c3e50')
        car_label.pack(pady=10)
        
        plate_label = tk.Label(dialog, text=f"Plate: {car_data['plate_number']}", 
                              font=('Arial', 11), bg='#f0f0f0', fg='#34495e')
        plate_label.pack()
        
        # Payment amount
        amount_label = tk.Label(dialog, text="Payment Amount (₱):", 
                               font=('Arial', 12), bg='#f0f0f0', fg='#34495e')
        amount_label.pack(pady=(20, 5))
        
        payment_entry = ttk.Entry(dialog, width=20, font=('Arial', 12))
        payment_entry.pack(pady=5)
        payment_entry.focus()
        
        # Buttons
        btn_frame = ttk.Frame(dialog)
        btn_frame.pack(pady=20)
        
        def confirm_payment():
            try:
                amount = float(payment_entry.get())
                if amount <= 0:
                    messagebox.showerror("Error", "Amount must be greater than 0")
                    return
                
                self.cars_data[car_id]['payment_amount'] = amount
                self.cars_data[car_id]['cashier_name'] = self.current_employee['name']
                self.cars_data[car_id]['status'] = self.STATUS_FINISHED
                self.cars_data[car_id]['completion_time'] = datetime.now()
                
                self.refresh_dashboard()
                dialog.destroy()
                messagebox.showinfo("Success", f"Payment of ₱{amount:.2f} processed successfully")
                
            except ValueError:
                messagebox.showerror("Error", "Please enter a valid amount")
        
        ttk.Button(btn_frame, text="Process Payment", 
                  command=confirm_payment).pack(side='left', padx=5)
        ttk.Button(btn_frame, text="Cancel", 
                  command=dialog.destroy).pack(side='left', padx=5)
        
        # Bind Enter key
        dialog.bind('<Return>', lambda e: confirm_payment())
    
    def show_add_car_dialog(self):
        """Show dialog to add a new car"""
        dialog = tk.Toplevel(self.root)
        dialog.title("Add New Car")
        dialog.geometry("450x350")
        dialog.transient(self.root)
        dialog.grab_set()
        dialog.configure(bg='#f0f0f0')
        
        # Car name
        car_name_label = tk.Label(dialog, text="Car Name:", 
                                 font=('Arial', 12), bg='#f0f0f0', fg='#34495e')
        car_name_label.pack(pady=5)
        car_name_entry = ttk.Entry(dialog, width=30, font=('Arial', 12))
        car_name_entry.pack(pady=5)
        car_name_entry.focus()
        
        # Plate number
        plate_label = tk.Label(dialog, text="Plate Number:", 
                              font=('Arial', 12), bg='#f0f0f0', fg='#34495e')
        plate_label.pack(pady=5)
        plate_entry = ttk.Entry(dialog, width=30, font=('Arial', 12))
        plate_entry.pack(pady=5)
        
        # Photo section
        photo_label = tk.Label(dialog, text="License Plate Photo (optional):", 
                              font=('Arial', 12), bg='#f0f0f0', fg='#34495e')
        photo_label.pack(pady=5)
        
        photo_frame = ttk.Frame(dialog)
        photo_frame.pack(pady=5)
        
        photo_path_var = tk.StringVar()
        photo_display = tk.Label(photo_frame, textvariable=photo_path_var, 
                                font=('Arial', 10), bg='#f0f0f0', fg='#3498db')
        photo_display.pack()
        
        def browse_photo():
            filename = filedialog.askopenfilename(
                title="Select License Plate Photo",
                filetypes=[("Image files", "*.jpg *.jpeg *.png *.gif *.bmp")]
            )
            if filename:
                photo_path_var.set(os.path.basename(filename))
                setattr(browse_photo, 'file_path', filename)
        
        ttk.Button(photo_frame, text="Browse Photo", 
                  command=browse_photo).pack(pady=5)
        
        # Buttons
        btn_frame = ttk.Frame(dialog)
        btn_frame.pack(pady=20)
        
        def add_car():
            car_name = car_name_entry.get().strip()
            plate_number = plate_entry.get().strip()
            
            if not car_name or not plate_number:
                messagebox.showerror("Error", "Please fill in all required fields")
                return
            
            # Generate unique ID
            car_id = f"car_{len(self.cars_data) + 1}_{datetime.now().strftime('%H%M%S')}"
            
            # Handle photo upload if selected
            photo_filename = None
            if hasattr(browse_photo, 'file_path'):
                try:
                    # Create uploads directory if it doesn't exist
                    uploads_dir = "uploads"
                    os.makedirs(uploads_dir, exist_ok=True)
                    
                    # Copy photo to uploads directory
                    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                    photo_filename = f"{timestamp}_{os.path.basename(browse_photo.file_path)}"
                    dest_path = os.path.join(uploads_dir, photo_filename)
                    shutil.copy2(browse_photo.file_path, dest_path)
                except Exception as e:
                    messagebox.showwarning("Warning", f"Failed to save photo: {e}")
            
            # Add car data
            self.cars_data[car_id] = {
                'car_name': car_name,
                'plate_number': plate_number,
                'status': self.STATUS_WASHING,
                'timestamp': datetime.now(),
                'washer_name': self.current_employee['name'],
                'cashier_name': None,
                'payment_amount': None,
                'completion_time': None,
                'photo_filename': photo_filename
            }
            
            self.refresh_dashboard()
            dialog.destroy()
            messagebox.showinfo("Success", f"Car '{car_name}' added to washing queue")
        
        ttk.Button(btn_frame, text="Add Car", 
                  command=add_car).pack(side='left', padx=5)
        ttk.Button(btn_frame, text="Cancel", 
                  command=dialog.destroy).pack(side='left', padx=5)
        
        # Bind Enter key
        dialog.bind('<Return>', lambda e: add_car())
    
    def export_excel(self):
        """Export daily data to Excel"""
        today = datetime.now().date()
        
        # Filter finished cars from today
        finished_today = []
        for car in self.cars_data.values():
            if (car['status'] == self.STATUS_FINISHED and 
                car.get('completion_time') and 
                car['completion_time'].date() == today):
                finished_today.append(car)
        
        if not finished_today:
            messagebox.showinfo("Info", "No completed cars found for today.")
            return
        
        # Ask for save location
        filename = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile=f"carwash_report_{today.strftime('%Y-%m-%d')}.xlsx"
        )
        
        if not filename:
            return
        
        try:
            # Create Excel workbook
            wb = Workbook()
            ws = wb.active
            ws.title = f"Carwash Report {today.strftime('%Y-%m-%d')}"
            
            # Define styles
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Headers
            headers = ['Car Name', 'Plate Number', 'Washer', 'Cashier', 'Payment Amount (₱)', 'Start Time', 'Completion Time']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
            
            # Data rows
            total_payment = 0
            for row, car in enumerate(finished_today, 2):
                data = [
                    car['car_name'],
                    car['plate_number'],
                    car.get('washer_name', ''),
                    car.get('cashier_name', ''),
                    car.get('payment_amount', 0),
                    car['timestamp'].strftime('%Y-%m-%d %H:%M:%S'),
                    car['completion_time'].strftime('%Y-%m-%d %H:%M:%S') if car.get('completion_time') else ''
                ]
                
                for col, value in enumerate(data, 1):
                    cell = ws.cell(row=row, column=col, value=value)
                    cell.border = border
                
                if car.get('payment_amount'):
                    total_payment += car['payment_amount']
            
            # Summary row
            summary_row = len(finished_today) + 3
            ws.cell(row=summary_row, column=1, value="TOTAL CARS:").font = Font(bold=True)
            ws.cell(row=summary_row, column=2, value=len(finished_today)).font = Font(bold=True)
            ws.cell(row=summary_row + 1, column=1, value="TOTAL REVENUE:").font = Font(bold=True)
            ws.cell(row=summary_row + 1, column=2, value=f"₱{total_payment:.2f}").font = Font(bold=True)
            
            # Auto-adjust column widths
            for col in range(1, len(headers) + 1):
                column_letter = get_column_letter(col)
                max_length = 0
                for row in range(1, len(finished_today) + 2):
                    cell_value = str(ws[f"{column_letter}{row}"].value)
                    max_length = max(max_length, len(cell_value))
                ws.column_dimensions[column_letter].width = min(max_length + 2, 30)
            
            # Save file
            wb.save(filename)
            messagebox.showinfo("Success", f"Excel report exported successfully!\nSaved as: {filename}")
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to export Excel: {e}")
    
    def export_csv(self):
        """Export daily data to CSV"""
        today = datetime.now().date()
        
        # Filter finished cars from today
        finished_today = []
        for car in self.cars_data.values():
            if (car['status'] == self.STATUS_FINISHED and 
                car.get('completion_time') and 
                car['completion_time'].date() == today):
                finished_today.append(car)
        
        if not finished_today:
            messagebox.showinfo("Info", "No completed cars found for today.")
            return
        
        # Ask for save location
        filename = filedialog.asksaveasfilename(
            defaultextension=".csv",
            filetypes=[("CSV files", "*.csv")],
            initialfile=f"carwash_report_{today.strftime('%Y-%m-%d')}.csv"
        )
        
        if not filename:
            return
        
        try:
            with open(filename, 'w', newline='', encoding='utf-8') as csvfile:
                writer = csv.writer(csvfile)
                
                # Write headers
                writer.writerow(['Car Name', 'Plate Number', 'Washer', 'Cashier', 
                               'Payment Amount (₱)', 'Start Time', 'Completion Time'])
                
                # Write data
                total_payment = 0
                for car in finished_today:
                    writer.writerow([
                        car['car_name'],
                        car['plate_number'],
                        car.get('washer_name', ''),
                        car.get('cashier_name', ''),
                        car.get('payment_amount', 0),
                        car['timestamp'].strftime('%Y-%m-%d %H:%M:%S'),
                        car['completion_time'].strftime('%Y-%m-%d %H:%M:%S') if car.get('completion_time') else ''
                    ])
                    
                    if car.get('payment_amount'):
                        total_payment += car['payment_amount']
                
                # Write summary
                writer.writerow([])
                writer.writerow(['TOTAL CARS:', len(finished_today)])
                writer.writerow(['TOTAL REVENUE:', f'₱{total_payment:.2f}'])
            
            messagebox.showinfo("Success", f"CSV report exported successfully!\nSaved as: {filename}")
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to export CSV: {e}")
    
    def reset_daily_data(self):
        """Reset all daily data"""
        total_cars = len(self.cars_data)
        finished_cars = len([car for car in self.cars_data.values() if car['status'] == self.STATUS_FINISHED])
        
        if total_cars == 0:
            messagebox.showinfo("Info", "No data to reset.")
            return
        
        # Confirmation dialog
        result = messagebox.askyesno(
            "Confirm Reset", 
            f"Are you sure you want to reset all daily data?\n\n"
            f"This will clear {total_cars} cars ({finished_cars} finished cars).\n"
            f"This action cannot be undone!"
        )
        
        if result:
            self.cars_data.clear()
            self.save_data()
            self.refresh_dashboard()
            messagebox.showinfo("Success", f"Daily data reset completed. Cleared {total_cars} cars.")
    
    def refresh_dashboard(self):
        """Refresh the dashboard display"""
        if hasattr(self, 'washing_tree'):
            self.populate_car_list(self.washing_tree, self.STATUS_WASHING)
        if hasattr(self, 'awaiting_payment_tree'):
            self.populate_car_list(self.awaiting_payment_tree, self.STATUS_AWAITING_PAYMENT)
        if hasattr(self, 'finished_tree'):
            self.populate_car_list(self.finished_tree, self.STATUS_FINISHED)
        
        self.update_status()
    
    def update_status(self):
        """Update status bar"""
        washing_count = len([c for c in self.cars_data.values() if c['status'] == self.STATUS_WASHING])
        payment_count = len([c for c in self.cars_data.values() if c['status'] == self.STATUS_AWAITING_PAYMENT])
        finished_count = len([c for c in self.cars_data.values() if c['status'] == self.STATUS_FINISHED])
        
        today_revenue = sum(c.get('payment_amount', 0) for c in self.cars_data.values() 
                           if c['status'] == self.STATUS_FINISHED and c.get('completion_time') 
                           and c['completion_time'].date() == datetime.now().date())
        
        status_text = f"Washing: {washing_count} | Awaiting Payment: {payment_count} | Finished: {finished_count} | Today's Revenue: ₱{today_revenue:.2f}"
        self.status_var.set(status_text)
    
    def auto_refresh(self):
        """Auto-refresh dashboard every 10 seconds"""
        self.refresh_dashboard()
        self.root.after(10000, self.auto_refresh)  # 10 seconds
    
    def logout(self):
        """Logout current employee"""
        self.current_employee = None
        self.save_data()
        self.create_login_screen()
    
    def clear_screen(self):
        """Clear all widgets from screen"""
        for widget in self.root.winfo_children():
            widget.destroy()

def main():
    """Main application entry point"""
    root = tk.Tk()
    
    # Configure style for better compatibility
    try:
        style = ttk.Style()
        style.theme_use('clam')
    except:
        # Use default theme if clam is not available
        pass
    
    app = CarwashApp(root)
    
    # Handle window closing
    def on_closing():
        app.save_data()
        root.destroy()
    
    root.protocol("WM_DELETE_WINDOW", on_closing)
    
    # Center window on screen
    root.update_idletasks()
    x = (root.winfo_screenwidth() // 2) - (root.winfo_width() // 2)
    y = (root.winfo_screenheight() // 2) - (root.winfo_height() // 2)
    root.geometry(f"+{x}+{y}")
    
    root.mainloop()

if __name__ == "__main__":
    main()